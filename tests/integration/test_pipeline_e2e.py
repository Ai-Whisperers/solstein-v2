"""End-to-end pipeline integration tests.

Runs the full `solstein run` pipeline against a small, deterministic universe,
with all network mocked. Validates the shape and content of the outputs:
Excel, Markdown brief, Narrative brief, Scored JSON.
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pytest_httpx import HTTPXMock

from solstein.domain import Company, Universe
from solstein.pipeline import run_pipeline


@pytest.fixture
def universe_with_manual_data() -> Universe:
    """Universe with enough pre-populated data to score honestly without network."""
    return Universe(
        name="integration-test",
        description="Integration test fixture for full pipeline.",
        companies=[
            # Scorable via manual data (growth + financial health)
            Company(
                name="HighGrower",
                country="NL",
                website="https://highgrower.example/",
                revenue_eur=50_000_000,
                employees=100,
                growth_yoy=0.45,
                founded_year=2015,
            ),
            Company(
                name="SteadyEddy",
                country="DE",
                website="https://steady.example/",
                revenue_eur=20_000_000,
                employees=80,
                growth_yoy=0.15,
                founded_year=2010,
            ),
            # Shrinking — should land in salt tier
            Company(
                name="Shrinking",
                country="FR",
                website="https://shrinking.example/",
                revenue_eur=10_000_000,
                employees=50,
                growth_yoy=-0.20,
                founded_year=2005,
            ),
            # Insufficient signal — should come back unknown
            Company(name="Mystery", country="ES"),
        ],
    )


def _mock_all_external_calls(httpx_mock: HTTPXMock) -> None:
    """Block / stub every external call a pipeline could make.

    We don't register specific responses — instead we allow any network request
    and return empty/failure, forcing adapters into their degrade-gracefully paths.
    This tests that the pipeline survives with no real external data.
    """
    # Allow all GitHub API calls to return empty (no github_org set on companies anyway)
    # and all website robots.txt / homepage to 404 → website adapter no-ops.
    httpx_mock.add_response(status_code=404, is_reusable=True)


class TestEndToEnd:
    @pytest.mark.asyncio
    async def test_pipeline_produces_all_four_outputs(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        out_dir = tmp_path / "out"
        result = await run_pipeline(universe_with_manual_data, out_dir)

        # Four output files
        assert (out_dir / f"{result.name}.xlsx").exists()
        assert (out_dir / f"{result.name}.md").exists()
        assert (out_dir / f"{result.name}-narrative.md").exists()
        assert (out_dir / f"{result.name}-scored.json").exists()

    @pytest.mark.asyncio
    async def test_pipeline_scores_companies_with_sufficient_data(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        result = await run_pipeline(universe_with_manual_data, tmp_path)

        by_name = {c.name: c for c in result.companies}
        assert by_name["HighGrower"].composite_score is not None
        assert by_name["SteadyEddy"].composite_score is not None
        assert by_name["Shrinking"].composite_score is not None
        assert by_name["Mystery"].composite_score is None
        assert by_name["Mystery"].tier == "unknown"

    @pytest.mark.asyncio
    async def test_ranking_order(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        result = await run_pipeline(universe_with_manual_data, tmp_path)

        # HighGrower (+45% YoY) should outrank SteadyEddy (+15%) outrank Shrinking (-20%)
        ranks = {c.name: i for i, c in enumerate(result.companies)}
        assert ranks["HighGrower"] < ranks["SteadyEddy"] < ranks["Shrinking"]
        # Mystery goes last (no score)
        assert ranks["Mystery"] == len(result.companies) - 1

    @pytest.mark.asyncio
    async def test_excel_output_has_ranking_sheet(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        result = await run_pipeline(universe_with_manual_data, tmp_path)

        wb = load_workbook(tmp_path / f"{result.name}.xlsx")
        assert "Ranking" in wb.sheetnames
        # First row is headers
        rows = list(wb["Ranking"].iter_rows(values_only=True))
        assert rows[0] == (
            "Rank",
            "Company",
            "Tier",
            "Composite",
            "Growth",
            "Financial Health",
            "AI Maturity",
            "Revenue (€)",
            "Employees",
            "YoY Growth",
            "GitHub Stars",
            "GitHub 90d Commits",
            "Completeness",
            "Country",
            "Website",
        )
        # 4 data rows
        assert len(rows) == 5

    @pytest.mark.asyncio
    async def test_scored_json_roundtrips(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        result = await run_pipeline(universe_with_manual_data, tmp_path)

        scored_path = tmp_path / f"{result.name}-scored.json"
        data = json.loads(scored_path.read_text())
        round_trip = Universe.model_validate(data)
        assert round_trip.name == result.name
        assert len(round_trip.companies) == len(result.companies)

    @pytest.mark.asyncio
    async def test_narrative_brief_mentions_universe_name(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
        universe_with_manual_data: Universe,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        result = await run_pipeline(universe_with_manual_data, tmp_path)

        narrative = (tmp_path / f"{result.name}-narrative.md").read_text()
        assert result.name in narrative
        assert "Executive observations" in narrative
        assert "Methodology note" in narrative


class TestDeterminism:
    """The pipeline must be deterministic with identical inputs."""

    @pytest.mark.asyncio
    async def test_two_runs_produce_identical_scored_json(
        self,
        tmp_path: Path,
        httpx_mock: HTTPXMock,
    ) -> None:
        _mock_all_external_calls(httpx_mock)

        u1 = Universe(
            name="det-1",
            companies=[
                Company(name="A", revenue_eur=30_000_000, employees=130, growth_yoy=0.22),
            ],
        )
        u2 = Universe(
            name="det-2",
            companies=[
                Company(name="A", revenue_eur=30_000_000, employees=130, growth_yoy=0.22),
            ],
        )

        r1 = await run_pipeline(u1, tmp_path / "r1")
        r2 = await run_pipeline(u2, tmp_path / "r2")

        assert r1.companies[0].composite_score == r2.companies[0].composite_score
        assert r1.companies[0].tier == r2.companies[0].tier


# Silence a false-positive unused-import warning when pytest_asyncio is installed
_ = asyncio
