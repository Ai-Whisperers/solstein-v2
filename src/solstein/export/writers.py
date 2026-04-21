"""Export writers: Excel shortlist + Markdown brief.

The deal team reads the Markdown. Excel is for anyone who wants to pivot.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from solstein.domain import Company, Tier, Universe

TIER_COLORS = {
    "phoenix": "C6EFCE",
    "diamond": "DDEBF7",
    "lead": "FFF2CC",
    "salt": "FCE4D6",
    "unknown": "FFFFFF",
}

_SHORTLIST_HEADERS = [
    "Rank",
    "Company",
    "Tier",
    "Composite",
    "Growth",
    "Financial Health",
    "AI Maturity",
    "Revenue (€)",
    "Employees",
    "YoY Growth",
    "GitHub Stars",
    "GitHub 90d Commits",
    "Completeness",
    "Country",
    "Website",
]

_COL_WIDTHS = {
    1: 6,  # Rank
    2: 28,  # Company
    3: 10,  # Tier
    4: 10,  # Composite
    5: 10,  # Growth
    6: 16,  # Financial Health
    7: 12,  # AI Maturity
    8: 14,  # Revenue
    9: 11,  # Employees
    10: 12,  # YoY Growth
    11: 12,  # GH Stars
    12: 18,  # GH Commits
    13: 14,  # Completeness
    14: 10,  # Country
    15: 40,  # Website
}


def _row_for(rank: int, c: Company) -> list[object]:
    return [
        rank,
        c.name,
        c.tier,
        c.composite_score,
        c.growth_score,
        c.financial_health_score,
        c.ai_maturity_score,
        c.revenue_eur,
        c.employees,
        c.growth_yoy,
        c.github_stars_total,
        c.github_commits_last_90d,
        round(c.completeness(), 2),
        c.country,
        str(c.website) if c.website else None,
    ]


def _style_sheet(ws: Worksheet, row_count: int) -> None:
    """Apply formatting: header bold + freeze, row shading, column widths."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if row_count > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_SHORTLIST_HEADERS))}{row_count + 1}"


def _fill_row(ws: Worksheet, row_idx: int, tier: Tier) -> None:
    fill = PatternFill(start_color=TIER_COLORS.get(tier, "FFFFFF"), fill_type="solid")
    for cell in ws[row_idx]:
        cell.fill = fill


def write_excel(universe: Universe, path: Path) -> None:
    """Write a multi-sheet workbook: Ranking, plus one sheet per non-empty tier."""
    wb = Workbook()

    # Main ranking sheet
    main = wb.active
    if main is None:
        raise RuntimeError("openpyxl gave no active sheet")
    main.title = "Ranking"

    main.append(_SHORTLIST_HEADERS)
    for rank, c in enumerate(universe.companies, start=1):
        main.append(_row_for(rank, c))
        _fill_row(main, main.max_row, c.tier)
    _style_sheet(main, len(universe.companies))

    # Per-tier sheets (only if the tier has members)
    for tier in ("phoenix", "diamond", "lead", "salt", "unknown"):
        members = [c for c in universe.companies if c.tier == tier]
        if not members:
            continue
        sheet = wb.create_sheet(title=tier.capitalize())
        sheet.append(_SHORTLIST_HEADERS)
        for rank, c in enumerate(members, start=1):
            sheet.append(_row_for(rank, c))
            _fill_row(sheet, sheet.max_row, c.tier)
        _style_sheet(sheet, len(members))

    wb.save(path)


def write_markdown_brief(universe: Universe, path: Path) -> None:
    lines: list[str] = [
        f"# {universe.name} — deal-team brief",
        "",
        f"_Generated {date.today().isoformat()} · {len(universe.companies)} companies._",
        "",
        "## Top candidates",
        "",
    ]
    top = [c for c in universe.companies if c.tier in ("phoenix", "diamond")][:10]
    if not top:
        lines.append("_No Phoenix or Diamond candidates in this universe._")
    else:
        for c in top:
            score = f"{c.composite_score:.2f}" if c.composite_score is not None else "n/a"
            lines.extend(
                [
                    f"### {c.name} — **{c.tier}** ({score}/10)",
                    "",
                    f"- Country: {c.country or 'unknown'}",
                    f"- Revenue: {_eur(c.revenue_eur)} · Employees: {c.employees or 'unknown'} · YoY: {_pct(c.growth_yoy)}",
                    f"- GitHub: {c.github_stars_total or 0} stars · {c.github_commits_last_90d or 0} commits (90d)",
                    f"- Completeness: {c.completeness():.0%}",
                    "",
                ]
            )

    lines.extend(
        [
            "",
            "## Full ranking",
            "",
            "| # | Company | Tier | Score | Completeness |",
            "|---|---|---|---|---|",
        ]
    )
    for rank, c in enumerate(universe.companies, start=1):
        score = f"{c.composite_score:.2f}" if c.composite_score is not None else "—"
        lines.append(f"| {rank} | {c.name} | {c.tier} | {score} | {c.completeness():.0%} |")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _eur(value: float | None) -> str:
    if value is None:
        return "unknown"
    if value >= 1_000_000:
        return f"€{value / 1_000_000:.1f}M"
    if value >= 1_000:
        return f"€{value / 1_000:.0f}K"
    return f"€{value:.0f}"


def _pct(value: float | None) -> str:
    return f"{value:.0%}" if value is not None else "unknown"
